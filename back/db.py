from openpyxl import load_workbook, Workbook
from os import mkdir, path, listdir
import random, string



class db():
    def __init__(self):
        if not path.exists("brands"):
            mkdir("brands")
        pass


    def generateHash(self, length):
        characters = string.ascii_lowercase + string.ascii_uppercase + string.digits
        return ''.join(random.choice(characters) for _ in range(length))


    def addBrand(self, brand):
        try:
            if path.exists("brands/" + brand):
                return (False, "Бренд уже существует")
            mkdir("brands/" + brand)
            return (True, "Успешно создан")
        except Exception as e:
            return (False, "Ошибка: " + str(e))


    def addModel(self, brand, model):
        try:
            if path.exists("brands/" + brand + model + ".xlsx"):
                return (False, "Модель уже существует")
            wb = Workbook()
            ws = wb.active
            ws.title = "Prices"
            ws.cell(row=1, column=1).value = "Качество"
            ws.cell(row=1, column=2).value = "Цена"
            ws.cell(row=1, column=3).value = "Количество"
            ws.cell(row=1, column=4).value = "ID"
            wb.save(f"brands/{brand}/" + model + ".xlsx")
            return (True, "Модель успешно созданна.")
        except Exception as e:
            return (False, "Ошибка: " + str(e))
        
    
    def addDisplay(self, brand, model, display, price, amount):
        try:
            namePath = f"brands/{brand}/{model}.xlsx"
            if not path.exists(namePath):
                return (False, "Модели не существует.")
            wb = load_workbook(namePath)
            ws = wb.active
            row = 2
            
            while not ws.cell(row=row, column=1).value == None: row += 1

            ws.cell(row=row, column=1).value = display
            ws.cell(row=row, column=2).value = price
            ws.cell(row=row, column=3).value = amount
            ws.cell(row=row, column=4).value = self.generateHash(8)

            wb.save(namePath)
            return (True, "Дисплей успешно добавлен.")
        except Exception as e:
            return (False, "Ошибка: " + str(e))


    def editDisplay(self, brand, model, display, price, amount, hash):
        try:
            namePath = f"brands/{brand}/{model}.xlsx"
            if not path.exists(namePath):
                return (False, "Модели не существует.")
            wb = load_workbook(namePath)
            ws = wb.active
            row = 2

            while ws.cell(row=row, column=4).value != hash and ws.cell(row=row, column=4).value != None: row += 1
            if ws.cell(row=row, column=4).value == None:
                return (False, "Ошибка при поиске не найден нужный хеш: " + hash)
            if ws.cell(row=row, column=4).value == hash:
                print(display, price, amount, hash)
                ws.cell(row=row, column=1).value = display
                ws.cell(row=row, column=2).value = price
                ws.cell(row=row, column=3).value = amount
                wb.save(namePath)
                return (True, "Успешно отредактирован хеш: " + hash)
        except Exception as e:
            return (False, "Ошибка: " + str(e))

    
    def getBrands(self):
        try:
            if not path.exists("brands"):
                return (False, "Брендов не существует.")
            brands = []
            for item in listdir("brands"):
                if path.isdir(path.join("brands", item)):
                    brands.append(item)
            return (True, brands)
        except Exception as e:
            return (False, "Ошибка: " + str(e))

        
    
    def getModels(self, brand):
        try:
            pathName = f"brands/{brand}"
            if not path.exists(pathName):
                return (False, "Бренд не найдены")
            
            models = []
            for item in listdir(pathName):
                full_path = path.join(pathName, item)
                if path.isfile(full_path):
                    models.append(item[:-5])
            return (True, models)
        except Exception as e:
            return (False, "Ошибка: " + str(e))
        
    
    def getDisplays(self, brand, model):
        try:
            pathName = f"brands/{brand}/{model}.xlsx"
            if not path.exists(pathName):
                return (False, "Бренд не найдены")
            
            row = 2
            displays = []
            wb = load_workbook(pathName)
            ws = wb.active

            while ws.cell(row=row, column=1).value != None:
                displays.append([ws.cell(row=row, column=1).value,
                                 ws.cell(row=row, column=2).value,
                                 ws.cell(row=row, column=3).value,
                                 ws.cell(row=row, column=4).value])
                row += 1
            return (True, displays)            
        except Exception as e:
            return (False, "Ошибка: " + str(e))